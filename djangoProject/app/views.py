import json
import csv
from datetime import timedelta, datetime

from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q, Count
from django.utils import timezone
from app.models import User, LoginStatistics, Data


def _parse_request_body(request):
    try:
        return json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return {}


@csrf_exempt
def login(request):
    if request.method != "POST":
        return JsonResponse({'message': '仅支持POST请求'}, status=405)

    data = _parse_request_body(request)
    identifier = data.get('identifier', '').strip()
    password = data.get('password', '').strip()

    if not all([identifier, password]):
        return JsonResponse({'message': '请填写账号和密码'}, status=400)

    user = User.objects.filter(
        Q(username=identifier) | Q(email=identifier)
    ).first()

    if not user:
        return JsonResponse({'message': '用户不存在'}, status=400)

    if password != user.password:
        return JsonResponse({'message': '密码错误'}, status=400)

    # 更新登录统计
    today = timezone.now().date()
    stats_obj, _ = LoginStatistics.objects.get_or_create(date=today)
    stats_obj.login_count = (stats_obj.login_count or 0) + 1
    stats_obj.save()

    # 登录成功后，把用户 id 写入 session
    request.session['user_id'] = user.id

    return JsonResponse({
        'message': '登录成功',
        'data': {
            'username': user.username,
            'email': user.email
        }
    })

@csrf_exempt
def register(request):
    if request.method != 'POST':
        return JsonResponse({'message': '仅支持POST请求'}, status=405)

    data = _parse_request_body(request)
    username = data.get('username', '').strip()
    email = data.get('email', '').strip()
    password = data.get('password', '').strip()
    confirm_password = data.get('confirmPassword', '').strip()

    if not all([username, email, password, confirm_password]):
        return JsonResponse({'message': '请完整填写用户名、邮箱、密码和确认密码'}, status=400)

    if password != confirm_password:
        return JsonResponse({'message': '两次输入的密码不一致'}, status=400)

    if User.objects.filter(username=username).exists():
        return JsonResponse({'message': '用户名已存在'}, status=400)

    if User.objects.filter(email=email).exists():
        return JsonResponse({'message': '邮箱已存在'}, status=400)

    # 创建用户（
    User.objects.create(
        username=username,
        email=email,
        password=password
    )

    return JsonResponse({'message': '注册成功'}, status=200)

@csrf_exempt
def logout(request):
    if request.method != 'POST':
        return JsonResponse({'message':'仅支持post请求'},status=405)

    request.session.flush()
    return JsonResponse({'message':'退出成功'})

# 仪表盘统计接口
def dashboard_stats(request):
    # 1. 顶部卡片统计
    user_count = User.objects.count()
    data_count = Data.objects.count()

    # 发货地址去重统计（城市数量）
    city_count = (
        Data.objects.exclude(ship_address__isnull=True)
        .exclude(ship_address='')
        .values('ship_address')
        .distinct()
        .count()
    )

    # 商品品牌去重统计（商品种类数量）
    product_count = (
        Data.objects.exclude(product_brand__isnull=True)
        .exclude(product_brand='')
        .values('product_brand')
        .distinct()
        .count()
    )

    # 店铺数量去重统计
    shop_count = (
        Data.objects.exclude(shop__isnull=True)
        .exclude(shop='')
        .values('shop')
        .distinct()
        .count()
    )

    # 2. 最近一周登录统计（折线图）
    today = timezone.now().date()
    start_date = today - timedelta(days=6)
    stats_qs = (
        LoginStatistics.objects.filter(date__gte=start_date, date__lte=today)
        .order_by('date')
        .values('date', 'login_count')
    )
    stats_map = {item['date']: item['login_count'] for item in stats_qs}
    login_stats = []
    for i in range(7):
        day = start_date + timedelta(days=i)
        login_stats.append({
            'date': day.strftime('%Y-%m-%d'),
            'count': int(stats_map.get(day, 0)),
        })

    # 3. 商品品牌分组统计（取前10，用于饼图/柱状图）
    product_group_qs = (
        Data.objects.exclude(product_brand__isnull=True)
        .exclude(product_brand='')
        .values('product_brand')
        .annotate(count=Count('product_brand'))
        .order_by('-count')[:10]
    )
    product_groups = [
        {'name': item['product_brand'], 'count': item['count']}
        for item in product_group_qs
    ]

    # 返回完整统计数据
    return JsonResponse({
        'stats': {
            'user_count': user_count,
            'data_count': data_count,
            'city_count': city_count,
            'product_count': product_count,
            'shop_count': shop_count,
        },
        'login_stats': login_stats,
        'product_groups': product_groups,
    })

def get_profile(request):
    """获取当前用户信息"""
    user_id = request.session.get('user_id')
    if not user_id:
        return JsonResponse({'message': '未登录'}, status=401)

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'message': '用户不存在'}, status=404)

    return JsonResponse({
        'message': '获取成功',
        'data': {
            'username': user.username,
            'email': user.email,
            'password': user.password,
            'user_text': user.user_text or ''
        }})


@csrf_exempt
def update_profile(request):
    """修改当前用户信息"""
    if request.method != 'POST':
        return JsonResponse({'message': '仅支持POST请求'}, status=405)

    user_id = request.session.get('user_id')
    if not user_id:
        return JsonResponse({'message': '未登录'}, status=401)

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'message': '用户不存在'}, status=404)

    data = _parse_request_body(request)
    username  = data.get('username', '').strip()
    email     = data.get('email', '').strip()
    user_text = data.get('user_text', '').strip()
    password  = data.get('password', '').strip()

    if not username or not email:
        return JsonResponse({'message': '用户名和邮箱不能为空'}, status=400)

    # 检查用户名是否被其他用户占用
    if User.objects.filter(username=username).exclude(id=user_id).exists():
        return JsonResponse({'message': '用户名已被占用'}, status=400)

    # 检查邮箱是否被其他用户占用
    if User.objects.filter(email=email).exclude(id=user_id).exists():
        return JsonResponse({'message': '邮箱已被占用'}, status=400)

    user.username  = username
    user.email     = email
    user.user_text = user_text
    if password:
        user.password = password
    user.save()

    return JsonResponse({'message': '保存成功'})


def get_filter_options(request):
    """获取筛选"""
    # 品牌去重
    brands = list(
        Data.objects.exclude(product_brand__isnull=True)
        .exclude(product_brand='')
        .values_list('product_brand', flat=True)
        .distinct()
        .order_by('product_brand')
    )

    # 地址去重
    addresses = list(
        Data.objects.exclude(ship_address__isnull=True)
        .exclude(ship_address='')
        .values_list('ship_address', flat=True)
        .distinct()
        .order_by('ship_address')
    )

    return JsonResponse({
        'brands': brands,
        'addresses': addresses
    })

def get_products(request):
    """获取商品列表"""
    # 获取筛选参数
    brands = request.GET.getlist('brands', [])
    addresses = request.GET.getlist('addresses', [])
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 10))

    # 构建查询
    queryset = Data.objects.all()

    if brands:
        queryset = queryset.filter(product_brand__in=brands)

    if addresses:
        queryset = queryset.filter(ship_address__in=addresses)

    # 总数
    total = queryset.count()

    # 分页
    start = (page - 1) * page_size
    end = start + page_size
    products = queryset[start:end]

    # 序列化数据
    product_list = []
    for product in products:
        product_list.append({
            'id': product.id,
            'title': product.title or '',
            'shop': product.shop or '',
            'price': float(product.price) if product.price else 0,
            'ship_address': product.ship_address or '',
            'image_url': product.image_url or '',
            'selling_points': product.selling_points or '',
            'detail_url': product.detail_url or '',
        })

    total_pages = (total + page_size - 1) // page_size

    return JsonResponse({
        'products': product_list,
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages
    })

@csrf_exempt
def export_products(request):
    """导出商品数据（Excel、CSV、TXT）"""
    if request.method != 'POST':
        return JsonResponse({'message': '仅支持POST请求'}, status=405)

    data = _parse_request_body(request)
    export_type = data.get('type', 'excel')  # excel, csv, txt
    brands = data.get('brands', [])
    addresses = data.get('addresses', [])

    # 构建查询
    queryset = Data.objects.all()

    if brands:
        queryset = queryset.filter(product_brand__in=brands)

    if addresses:
        queryset = queryset.filter(ship_address__in=addresses)

    products = queryset.values(
        'title', 'shop', 'price', 'pay_count', 'ship_address',
        'detail_url', 'image_url', 'features', 'selling_points', 'product_brand'
    )

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Excel 导出逻辑
    if export_type == 'excel':
        try:
            import openpyxl
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "商品数据"

            # 表头
            headers = ['标题', '店铺', '价格', '付款人数', '发货地址', '详情页地址', '图片地址', '特点列表', '卖点', '商品品牌']
            ws.append(headers)

            # 数据
            for product in products:
                ws.append([
                    product.get('title', ''),
                    product.get('shop', ''),
                    product.get('price', ''),
                    product.get('pay_count', ''),
                    product.get('ship_address', ''),
                    product.get('detail_url', ''),
                    product.get('image_url', ''),
                    product.get('features', ''),
                    product.get('selling_points', ''),
                    product.get('product_brand', ''),
                ])

            # 构造响应
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="商品数据_{timestamp}.xlsx"'
            wb.save(response)
            return response

        except ImportError:
            return JsonResponse({'message': '需要安装openpyxl库: pip install openpyxl'}, status=500)

    # CSV 导出逻辑
    elif export_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="商品数据_{timestamp}.csv"'

        writer = csv.writer(response)
        writer.writerow(['标题', '店铺', '价格', '付款人数', '发货地址', '详情页地址', '图片地址', '特点列表', '卖点', '商品品牌'])

        for product in products:
            writer.writerow([
                product.get('title', ''),
                product.get('shop', ''),
                product.get('price', ''),
                product.get('pay_count', ''),
                product.get('ship_address', ''),
                product.get('detail_url', ''),
                product.get('image_url', ''),
                product.get('features', ''),
                product.get('selling_points', ''),
                product.get('product_brand', ''),
            ])
        return response

    # TXT 导出逻辑
    elif export_type == 'txt':
        response = HttpResponse(content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="商品数据_{timestamp}.txt"'

        lines = []
        lines.append('商品数据导出\n')
        lines.append('=' * 80 + '\n\n')

        for idx, product in enumerate(products, 1):
            lines.append(f'商品 {idx}:\n')
            lines.append(f'  标题: {product.get("title", "")}\n')
            lines.append(f'  店铺: {product.get("shop", "")}\n')
            lines.append(f'  价格: {product.get("price", "")}\n')
            lines.append(f'  付款人数: {product.get("pay_count", "")}\n')
            lines.append(f'  发货地址: {product.get("ship_address", "")}\n')
            lines.append(f'  详情页地址: {product.get("detail_url", "")}\n')
            lines.append(f'  图片地址: {product.get("image_url", "")}\n')
            lines.append(f'  特点列表: {product.get("features", "")}\n')
            lines.append(f'  卖点: {product.get("selling_points", "")}\n')
            lines.append(f'  商品品牌: {product.get("product_brand", "")}\n')
            lines.append('-' * 80 + '\n\n')

        response.write(''.join(lines))
        return response

    # 不支持的导出类型
    else:
        return JsonResponse({'message': '不支持的导出类型'}, status=400)


from .until import qurey

def _build_part1_dict(rows):
    """将 part1 查询结果聚合为 {name: value} 字典"""
    data_dict = {}
    for row in rows:
        name = row[0].strip() if row[0] else ''
        try:
            value = float(row[1]) if row[1] is not None else 0
        except (ValueError, TypeError):
            value = 0
        if name:
            if name in data_dict:
                data_dict[name] += value
            else:
                data_dict[name] = value
    return data_dict


def get_part1_data(request):
    """获取 part1 全量数据"""
    try:
        rows = qurey("SELECT name, value FROM part1")
        data_dict = _build_part1_dict(rows)
        data_list = [{'name': k, 'value': v} for k, v in data_dict.items()]
        names = sorted(data_dict.keys())
        return JsonResponse({'data': data_list, 'names': names})
    except Exception as e:
        return JsonResponse({'message': f'获取失败:{str(e)}'}, status=500)


def get_part1_filtered(request):
    """按名称筛选 part1 数据"""
    try:
        selected = request.GET.getlist('names', [])
        if selected:
            placeholders = ','.join(['%s'] * len(selected))
            rows = qurey(f"SELECT name, value FROM part1 WHERE name IN ({placeholders})" % tuple(f"'{n}'" for n in selected))
        else:
            rows = qurey("SELECT name, value FROM part1")
        data_dict = _build_part1_dict(rows)
        data_list = [{'name': k, 'value': v} for k, v in data_dict.items()]
        return JsonResponse({'data': data_list})
    except Exception as e:
        return JsonResponse({'message': f'筛选失败:{str(e)}'}, status=500)


@csrf_exempt
def export_part1(request):
    """导出 part1 数据（Excel、CSV、TXT）"""
    if request.method != 'POST':
        return JsonResponse({'message': '仅支持POST请求'}, status=405)

    data = _parse_request_body(request)
    export_type = data.get('type', 'excel')
    selected = data.get('names', [])

    try:
        if selected:
            placeholders = ','.join([f"'{n}'" for n in selected])
            rows = qurey(f"SELECT name, value FROM part1 WHERE name IN ({placeholders})")
        else:
            rows = qurey("SELECT name, value FROM part1")
    except Exception as e:
        return JsonResponse({'message': f'查询失败:{str(e)}'}, status=500)

    data_dict = _build_part1_dict(rows)
    data_list = [{'name': k, 'value': v} for k, v in data_dict.items()]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if export_type == 'excel':
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "part1数据"
            ws.append(['名称', '数值'])
            for item in data_list:
                ws.append([item['name'], item['value']])
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="part1_{timestamp}.xlsx"'
            wb.save(response)
            return response
        except ImportError:
            return JsonResponse({'message': '需要安装openpyxl: pip install openpyxl'}, status=500)

    elif export_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="part1_{timestamp}.csv"'
        writer = csv.writer(response)
        writer.writerow(['名称', '数值'])
        for item in data_list:
            writer.writerow([item['name'], item['value']])
        return response

    elif export_type == 'txt':
        response = HttpResponse(content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="part1_{timestamp}.txt"'
        lines = ['part1数据导出\n', '=' * 40 + '\n\n']
        for idx, item in enumerate(data_list, 1):
            lines.append(f'{idx}. {item["name"]}: {item["value"]}\n')
        response.write(''.join(lines))
        return response

    else:
        return JsonResponse({'message': '不支持的导出类型'}, status=400)


def _get_part_data(table, request_names=None):
    """通用：查询 part 表数据，可选按 names 筛选"""
    if request_names:
        placeholders = ','.join([f"'{n}'" for n in request_names])
        rows = qurey(f"SELECT name, value FROM {table} WHERE name IN ({placeholders})")
    else:
        rows = qurey(f"SELECT name, value FROM {table}")
    return _build_part1_dict(rows)


def get_part3_data(request):
    """获取 part3 全量数据"""
    try:
        data_dict = _get_part_data('part3')
        data_list = [{'name': k, 'value': v} for k, v in data_dict.items()]
        names = sorted(data_dict.keys())
        return JsonResponse({'data': data_list, 'names': names})
    except Exception as e:
        return JsonResponse({'message': f'获取失败:{str(e)}'}, status=500)


def get_part3_filtered(request):
    """按名称筛选 part3 数据"""
    try:
        selected = request.GET.getlist('names', [])
        data_dict = _get_part_data('part3', selected if selected else None)
        data_list = [{'name': k, 'value': v} for k, v in data_dict.items()]
        return JsonResponse({'data': data_list})
    except Exception as e:
        return JsonResponse({'message': f'筛选失败:{str(e)}'}, status=500)


@csrf_exempt
def export_part3(request):
    """导出 part3 数据（Excel、CSV、TXT）"""
    if request.method != 'POST':
        return JsonResponse({'message': '仅支持POST请求'}, status=405)

    data = _parse_request_body(request)
    export_type = data.get('type', 'excel')
    selected = data.get('names', [])

    try:
        data_dict = _get_part_data('part3', selected if selected else None)
    except Exception as e:
        return JsonResponse({'message': f'查询失败:{str(e)}'}, status=500)

    data_list = [{'name': k, 'value': v} for k, v in data_dict.items()]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if export_type == 'excel':
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "店铺销量分析"
            ws.append(['店铺名称', '销量'])
            for item in data_list:
                ws.append([item['name'], item['value']])
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="店铺销量分析_{timestamp}.xlsx"'
            wb.save(response)
            return response
        except ImportError:
            return JsonResponse({'message': '需要安装openpyxl: pip install openpyxl'}, status=500)

    elif export_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="店铺销量分析_{timestamp}.csv"'
        writer = csv.writer(response)
        writer.writerow(['店铺名称', '销量'])
        for item in data_list:
            writer.writerow([item['name'], item['value']])
        return response

    elif export_type == 'txt':
        response = HttpResponse(content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="店铺销量分析_{timestamp}.txt"'
        lines = ['店铺销量分析导出\n', '=' * 40 + '\n\n']
        for idx, item in enumerate(data_list, 1):
            lines.append(f'{idx}. {item["name"]}: {item["value"]}\n')
        response.write(''.join(lines))
        return response

    else:
        return JsonResponse({'message': '不支持的导出类型'}, status=400)

def _map_province_name(name):
    """将省名全称映射为完整名称（省、市、自治区等后缀），用于地图匹配"""
    if not name:
        return name

    name = name.strip()

    # 已包含后缀的直接返回
    if name.endswith(('省', '市', '自治区', '特别行政区')):
        return name

    # 省级行政区映射
    province_mapping = {
        # 直辖市
        '北京': '北京市',
        '天津': '天津市',
        '上海': '上海市',
        '重庆': '重庆市',

        # 自治区
        '新疆': '新疆维吾尔自治区',
        '西藏': '西藏自治区',
        '内蒙古': '内蒙古自治区',
        '宁夏': '宁夏回族自治区',
        '广西': '广西壮族自治区',

        # 特别行政区
        '香港': '香港特别行政区',
        '澳门': '澳门特别行政区',
        '台湾': '台湾省',

        # 省份（添加"省"后缀）
        '河北': '河北省',
        '山西': '山西省',
        '辽宁': '辽宁省',
        '吉林': '吉林省',
        '黑龙江': '黑龙江省',
        '江苏': '江苏省',
        '浙江': '浙江省',
        '安徽': '安徽省',
        '福建': '福建省',
        '江西': '江西省',
        '山东': '山东省',
        '河南': '河南省',
        '湖北': '湖北省',
        '湖南': '湖南省',
        '广东': '广东省',
        '海南': '海南省',
        '四川': '四川省',
        '贵州': '贵州省',
        '云南': '云南省',
        '陕西': '陕西省',
        '甘肃': '甘肃省',
        '青海': '青海省',
    }

    # 如果存在映射关系，返回映射后的名称
    if name in province_mapping:
        return province_mapping[name]

    # 默认添加"省"后缀
    return f"{name}省"

def get_part2_data(request):
    """获取part2表数据（城市销售额分析）"""
    try:
        # 使用工具类查询数据库
        sql = "SELECT name, value FROM part2"
        results = qurey(sql)

        # 处理数据：按name去重并累加value
        data_dict = {}
        for row in results:
            name = row[0].strip() if row[0] else ''
            value = float(row[1]) if row[1] else 0
            if name:
                if name in data_dict:
                    data_dict[name] += value
                else:
                    data_dict[name] = value

        # 转换为列表格式，将短名称映射为完整名称
        data_list = []
        for original_name, value in data_dict.items():
            mapped_name = _map_province_name(original_name)  # 映射为完整名称
            print(mapped_name)
            data_list.append({
                'name': mapped_name,  # 完整名称，用于地图匹配和显示
                'value': value
            })

        # 获取所有去重后的原始name列表（用于筛选）
        names = sorted(list(data_dict.keys()))

        return JsonResponse({
            'data': data_list,
            'names': names
        })
    except Exception as e:
        return JsonResponse({'message': f'获取数据失败: {str(e)}'}, status=500)


def get_part2_filtered(request):
    """按名称筛选 part2 数据（返回映射后的完整省名）"""
    try:
        selected = request.GET.getlist('names', [])
        if selected:
            placeholders = ','.join([f"'{n}'" for n in selected])
            rows = qurey(f"SELECT name, value FROM part2 WHERE name IN ({placeholders})")
        else:
            rows = qurey("SELECT name, value FROM part2")

        data_dict = {}
        for row in rows:
            name = row[0].strip() if row[0] else ''
            value = float(row[1]) if row[1] else 0
            if name:
                data_dict[name] = data_dict.get(name, 0) + value

        data_list = [{'name': _map_province_name(k), 'value': v} for k, v in data_dict.items()]
        return JsonResponse({'data': data_list})
    except Exception as e:
        return JsonResponse({'message': f'筛选失败: {str(e)}'}, status=500)


@csrf_exempt
def export_part2(request):
    """导出 part2 数据（Excel、CSV、TXT）"""
    if request.method != 'POST':
        return JsonResponse({'message': '仅支持POST请求'}, status=405)

    data = _parse_request_body(request)
    export_type = data.get('type', 'excel')
    selected = data.get('names', [])

    try:
        if selected:
            placeholders = ','.join([f"'{n}'" for n in selected])
            rows = qurey(f"SELECT name, value FROM part2 WHERE name IN ({placeholders})")
        else:
            rows = qurey("SELECT name, value FROM part2")
    except Exception as e:
        return JsonResponse({'message': f'查询失败: {str(e)}'}, status=500)

    data_dict = {}
    for row in rows:
        name = row[0].strip() if row[0] else ''
        value = float(row[1]) if row[1] else 0
        if name:
            data_dict[name] = data_dict.get(name, 0) + value

    data_list = [{'name': _map_province_name(k), 'value': v} for k, v in data_dict.items()]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if export_type == 'excel':
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "城市销售额分析"
            ws.append(['省份', '销售额'])
            for item in data_list:
                ws.append([item['name'], item['value']])
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="城市销售额分析_{timestamp}.xlsx"'
            wb.save(response)
            return response
        except ImportError:
            return JsonResponse({'message': '需要安装openpyxl: pip install openpyxl'}, status=500)

    elif export_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="城市销售额分析_{timestamp}.csv"'
        writer = csv.writer(response)
        writer.writerow(['省份', '销售额'])
        for item in data_list:
            writer.writerow([item['name'], item['value']])
        return response

    elif export_type == 'txt':
        response = HttpResponse(content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="城市销售额分析_{timestamp}.txt"'
        lines = ['城市销售额分析导出\n', '=' * 40 + '\n\n']
        for idx, item in enumerate(data_list, 1):
            lines.append(f'{idx}. {item["name"]}: {item["value"]}\n')
        response.write(''.join(lines))
        return response

    else:
        return JsonResponse({'message': '不支持的导出类型'}, status=400)


def _make_export_response(data_list, export_type, filename_prefix, col_headers):
    """通用导出响应生成器"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    keys = list(data_list[0].keys()) if data_list else []
    if export_type == 'excel':
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = filename_prefix
            ws.append(col_headers)
            for item in data_list:
                ws.append([item.get(k, '') for k in keys])
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timestamp}.xlsx"'
            wb.save(response)
            return response
        except ImportError:
            return JsonResponse({'message': '需要安装openpyxl: pip install openpyxl'}, status=500)
    elif export_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timestamp}.csv"'
        writer = csv.writer(response)
        writer.writerow(col_headers)
        for item in data_list:
            writer.writerow([item.get(k, '') for k in keys])
        return response
    elif export_type == 'txt':
        response = HttpResponse(content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timestamp}.txt"'
        lines = [f'{filename_prefix}导出\n', '=' * 40 + '\n\n']
        for idx, item in enumerate(data_list, 1):
            row = ' | '.join(str(item.get(k, '')) for k in keys)
            lines.append(f'{idx}. {row}\n')
        response.write(''.join(lines))
        return response
    else:
        return JsonResponse({'message': '不支持的导出类型'}, status=400)


# ── part4 ──────────────────────────────────────────────
def get_part4_data(request):
    try:
        data_dict = _get_part_data('part4')
        data_list = [{'name': k, 'value': v} for k, v in data_dict.items()]
        return JsonResponse({'data': data_list, 'names': sorted(data_dict.keys())})
    except Exception as e:
        return JsonResponse({'message': f'获取失败:{str(e)}'}, status=500)

def get_part4_filtered(request):
    try:
        selected = request.GET.getlist('names', [])
        data_dict = _get_part_data('part4', selected or None)
        return JsonResponse({'data': [{'name': k, 'value': v} for k, v in data_dict.items()]})
    except Exception as e:
        return JsonResponse({'message': f'筛选失败:{str(e)}'}, status=500)

@csrf_exempt
def export_part4(request):
    if request.method != 'POST':
        return JsonResponse({'message': '仅支持POST请求'}, status=405)
    data = _parse_request_body(request)
    selected = data.get('names', [])
    try:
        data_dict = _get_part_data('part4', selected or None)
    except Exception as e:
        return JsonResponse({'message': f'查询失败:{str(e)}'}, status=500)
    data_list = [{'name': k, 'value': v} for k, v in data_dict.items()]
    return _make_export_response(data_list, data.get('type', 'excel'), '商品关键词分析', ['关键词', '数量'])


# ── part5 ──────────────────────────────────────────────
def get_part5_data(request):
    try:
        rows = qurey("SELECT name, value FROM part5")
        data_list = []
        for row in rows:
            name = row[0].strip() if row[0] else ''
            try:
                value = float(row[1]) if row[1] is not None else 0
            except (ValueError, TypeError):
                value = 0
            if name:
                data_list.append({'name': name, 'value': value})
        return JsonResponse({'data': data_list})
    except Exception as e:
        return JsonResponse({'message': f'获取失败:{str(e)}'}, status=500)

@csrf_exempt
def export_part5(request):
    if request.method != 'POST':
        return JsonResponse({'message': '仅支持POST请求'}, status=405)
    data = _parse_request_body(request)
    try:
        rows = qurey("SELECT name, value FROM part5")
    except Exception as e:
        return JsonResponse({'message': f'查询失败:{str(e)}'}, status=500)
    data_list = []
    for row in rows:
        name = row[0].strip() if row[0] else ''
        try:
            value = float(row[1]) if row[1] is not None else 0
        except (ValueError, TypeError):
            value = 0
        if name:
            data_list.append({'name': name, 'value': value})
    return _make_export_response(data_list, data.get('type', 'excel'), '商品特点分布', ['特点', '数量'])


# ── part6 ──────────────────────────────────────────────
def get_part6_data(request):
    try:
        data_dict = _get_part_data('part6')
        data_list = [{'name': k, 'value': v} for k, v in data_dict.items()]
        return JsonResponse({'data': data_list, 'names': sorted(data_dict.keys())})
    except Exception as e:
        return JsonResponse({'message': f'获取失败:{str(e)}'}, status=500)

@csrf_exempt
def export_part6(request):
    if request.method != 'POST':
        return JsonResponse({'message': '仅支持POST请求'}, status=405)
    data = _parse_request_body(request)
    try:
        data_dict = _get_part_data('part6')
    except Exception as e:
        return JsonResponse({'message': f'查询失败:{str(e)}'}, status=500)
    data_list = [{'name': k, 'value': v} for k, v in data_dict.items()]
    return _make_export_response(data_list, data.get('type', 'excel'), '价格区间分析', ['价格区间', '销售额'])


# ── part7 ──────────────────────────────────────────────
def get_part7_data(request):
    try:
        data_dict = _get_part_data('part7')
        data_list = [{'name': k, 'value': v} for k, v in data_dict.items()]
        return JsonResponse({'data': data_list, 'names': sorted(data_dict.keys())})
    except Exception as e:
        return JsonResponse({'message': f'获取失败:{str(e)}'}, status=500)

def get_part7_filtered(request):
    try:
        selected = request.GET.getlist('names', [])
        data_dict = _get_part_data('part7', selected or None)
        return JsonResponse({'data': [{'name': k, 'value': v} for k, v in data_dict.items()]})
    except Exception as e:
        return JsonResponse({'message': f'筛选失败:{str(e)}'}, status=500)

@csrf_exempt
def export_part7(request):
    if request.method != 'POST':
        return JsonResponse({'message': '仅支持POST请求'}, status=405)
    data = _parse_request_body(request)
    selected = data.get('names', [])
    try:
        data_dict = _get_part_data('part7', selected or None)
    except Exception as e:
        return JsonResponse({'message': f'查询失败:{str(e)}'}, status=500)
    data_list = [{'name': k, 'value': v} for k, v in data_dict.items()]
    return _make_export_response(data_list, data.get('type', 'excel'), '商品评分分析', ['商品', '评分'])


# ── 销量预测 ──────────────────────────────────────────────
import os, joblib
import numpy as np

_RF_BUNDLE = None
_RF_BUNDLE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '..', 'spider', 'models', 'rf_model_bundle.joblib'
)

def _load_rf_bundle():
    global _RF_BUNDLE
    if _RF_BUNDLE is None:
        _RF_BUNDLE = joblib.load(_RF_BUNDLE_PATH)
    return _RF_BUNDLE


def _safe_encode(encoder, value):
    """对未见过的标签返回 0，避免 LabelEncoder 报错"""
    try:
        return int(encoder.transform([str(value)])[0])
    except ValueError:
        return 0


@csrf_exempt
def predict_pay_count(request):
    if request.method != 'POST':
        return JsonResponse({'message': '仅支持POST请求'}, status=405)

    data = _parse_request_body(request)
    title          = data.get('title', '')
    shop           = data.get('shop', '')
    price          = float(data.get('price', 0) or 0)
    ship_address   = data.get('ship_address', '')
    features       = data.get('features', '')
    product_brand  = data.get('product_brand', '')

    try:
        bundle   = _load_rf_bundle()
        model    = bundle['model']
        encoders = bundle['model_encoders']

        title_len      = len(str(title))
        features_count = str(features).count(',') + 1 if features else 0
        shop_code      = _safe_encode(encoders['shop_code'],    shop or '未知')
        brand_code     = _safe_encode(encoders['brand_code'],   product_brand or '未知')
        address_code   = _safe_encode(encoders['address_code'], ship_address or '未知')

        X = np.array([[price, title_len, features_count, shop_code, brand_code, address_code]])
        prediction = int(round(float(model.predict(X)[0])))
        prediction = max(0, prediction)

        return JsonResponse({
            'prediction': prediction,
            'model_info': {
                'algorithm': '随机森林回归 (Random Forest Regressor)',
                'training_samples': len(model.estimators_[0].tree_.value) if hasattr(model, 'estimators_') else 0,
                'features_used': ['售价', '标题长度', '特点数量', '店铺', '品牌', '发货地址'],
            }
        })
    except Exception as e:
        return JsonResponse({'message': f'预测失败: {str(e)}'}, status=500)


# ── 市场聚类分析 ──────────────────────────────────────────────
import csv as _csv

_KMEANS_CSV_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '..', 'spider', 'models', 'kmeans_results.csv'
)

def cluster_analysis(request):
    try:
        clusters_dict = {}
        with open(_KMEANS_CSV_PATH, encoding='utf-8-sig') as f:
            reader = _csv.DictReader(f)
            for row in reader:
                cid   = int(row['cluster'])
                cname = row['cluster_name']
                try:
                    price     = float(row['price'])
                    pay_count = int(float(row['pay_count']))
                except (ValueError, TypeError):
                    continue
                title = row.get('title', '')

                if cid not in clusters_dict:
                    clusters_dict[cid] = {'cluster_id': cid, 'name': cname, 'points': []}
                clusters_dict[cid]['points'].append({
                    'price':     price,
                    'pay_count': pay_count,
                    'title':     title,
                })

        clusters = []
        for cid in sorted(clusters_dict.keys()):
            c = clusters_dict[cid]
            c['count'] = len(c['points'])
            clusters.append(c)

        return JsonResponse({'clusters': clusters})
    except FileNotFoundError:
        return JsonResponse({'message': 'kmeans_results.csv 不存在，请先运行 tran_kmeans.py 生成聚类结果'}, status=500)
    except Exception as e:
        return JsonResponse({'message': f'聚类数据读取失败: {str(e)}'}, status=500)

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import Data  # 需确保Data模型已正确导入

@csrf_exempt
def get_product_detail(request, product_id):
    """获取商品详情"""
    try:
        product = Data.objects.get(id=product_id)
        return JsonResponse({
            'message': '获取成功',
            'data': {
                'id': product.id,
                'title': product.title or '',
                'shop': product.shop or '',
                'price': float(product.price) if product.price else 0,
                'pay_count': product.pay_count or 0,
                'ship_address': product.ship_address or '',
                'detail_url': product.detail_url or '',
                'image_url': product.image_url or '',
                'features': product.features or '',
                'selling_points': product.selling_points or '',
                'product_brand': product.product_brand or '',
            }
        })
    except Data.DoesNotExist:
        return JsonResponse({'message': '商品不存在'}, status=404)
    except Exception as e:
        return JsonResponse({'message': f'获取详情失败: {str(e)}'}, status=500)

@csrf_exempt
def chat_deepseek(request):
    """在线客服"""
    if request.method != 'POST':
        return JsonResponse({'message': '仅支持POST请求'}, status=405)

    data = _parse_request_body(request)
    user_message = data.get('message', '').strip()

    if not user_message:
        return JsonResponse({'message': '请输入要咨询的内容'}, status=400)

    api_key = "sk-2e09e7ac0dec41dba1c4507eaa06c170"
    if not api_key:
        return JsonResponse({'message': '你还没有配置Key或者已经失效'}, status=500)

    try:
        from openai import OpenAI

        client = OpenAI(
            api_key=api_key,
            base_url="https://api.deepseek.com"
        )

        completion = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "你是淘宝商品在线客服, 请用简洁中文回答, 控制在100字以内。"},
                {"role": "user", "content": user_message},
            ],
            stream=False
        )

        reply = completion.choices[0].message.content if completion and completion.choices else ''
        return JsonResponse({'reply': reply})
    except Exception as e:
        return JsonResponse({'message': f'调用失败: {str(e)}'}, status=500)


